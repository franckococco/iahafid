"""Excel de fichas: una fila = una pieza. El mismo auto+conjunto arma una ficha."""

from __future__ import annotations

from collections import OrderedDict
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from taller.knowledge import fold, identity_key, load_all, upsert

COLUMNS = (
    ("marca", "Marca del auto"),
    ("modelo", "Modelo"),
    ("motor", "Motor"),
    ("anio", "Año"),
    ("conjunto", "Conjunto (tren delantero, embrague…)"),
    ("notas", "Notas del local"),
    ("venta", "Cómo vender / cruzar"),
    ("pieza", "Nombre de la pieza"),
    ("marca_pieza", "Marca de la pieza (Sachs, etc.)"),
    ("lado", "Lado / par"),
    ("codigo", "Código"),
    ("nota", "Nota de la pieza"),
)

_ALIASES = {
    "marca": {"marca"},
    "modelo": {"modelo"},
    "motor": {"motor", "cilindrada"},
    "anio": {"anio", "ano", "year"},
    "conjunto": {"conjunto", "sistema", "grupo"},
    "notas": {"notas", "notas_del_local", "notas_local"},
    "venta": {"venta", "como_vender", "cruzar", "como_vender_cruzar"},
    "pieza": {"pieza", "nombre", "repuesto", "nombre_pieza"},
    "marca_pieza": {"marca_pieza", "marca_del_repuesto", "marca_repuesto"},
    "lado": {"lado", "posicion"},
    "codigo": {"codigo", "oem"},
    "nota": {"nota", "nota_pieza"},
    "id": {"id", "ficha_id"},
}

_EXAMPLES = (
    {
        "marca": "volkswagen",
        "modelo": "gol trend",
        "motor": "1.4",
        "anio": "2014",
        "conjunto": "tren delantero",
        "notas": "De a pares. No mezclar con 1.6.",
        "venta": "Si pide cazoleta o rótula, ofrecé amortiguador Sachs y bujes. El Sachs le va bien a este 1.4.",
        "pieza": "Cazoleta",
        "marca_pieza": "",
        "lado": "par",
        "codigo": "",
        "nota": "",
    },
    {
        "marca": "volkswagen",
        "modelo": "gol trend",
        "motor": "1.4",
        "anio": "2014",
        "conjunto": "tren delantero",
        "notas": "De a pares. No mezclar con 1.6.",
        "venta": "Si pide cazoleta o rótula, ofrecé amortiguador Sachs y bujes. El Sachs le va bien a este 1.4.",
        "pieza": "Amortiguador delantero",
        "marca_pieza": "Sachs",
        "lado": "par",
        "codigo": "",
        "nota": "1.4 nafta",
    },
    {
        "marca": "volkswagen",
        "modelo": "gol trend",
        "motor": "1.4",
        "anio": "2014",
        "conjunto": "tren delantero",
        "notas": "De a pares. No mezclar con 1.6.",
        "venta": "Si pide cazoleta o rótula, ofrecé amortiguador Sachs y bujes. El Sachs le va bien a este 1.4.",
        "pieza": "Bujes de tren",
        "marca_pieza": "",
        "lado": "par",
        "codigo": "",
        "nota": "",
    },
    {
        "marca": "volkswagen",
        "modelo": "gol trend",
        "motor": "1.4",
        "anio": "2014",
        "conjunto": "tren delantero",
        "notas": "De a pares. No mezclar con 1.6.",
        "venta": "Si pide cazoleta o rótula, ofrecé amortiguador Sachs y bujes. El Sachs le va bien a este 1.4.",
        "pieza": "Rótula inferior",
        "marca_pieza": "",
        "lado": "par",
        "codigo": "",
        "nota": "Confirmar lado al pedir",
    },
)


class ExcelError(ValueError):
    pass


def export_bytes(fichas: list[dict] | None = None) -> bytes:
    rows = _rows_from_fichas(fichas if fichas is not None else load_all())
    if not rows:
        rows = list(_EXAMPLES)
    book = Workbook()
    sheet = book.active
    sheet.title = "fichas"
    _write_sheet(sheet, rows)
    leeme = book.create_sheet("LEEME")
    for index, line in enumerate(_LEEME, start=1):
        leeme.cell(index, 1, line)
        leeme.cell(index, 1).alignment = Alignment(wrap_text=True)
    leeme.column_dimensions["A"].width = 92
    buf = BytesIO()
    book.save(buf)
    return buf.getvalue()


def import_bytes(raw: bytes) -> dict:
    if not raw:
        raise ExcelError("El archivo está vacío.")
    try:
        book = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelError("No pude leer el Excel. ¿Es un .xlsx?") from exc
    sheet = _data_sheet(book)
    headers, start = _find_headers(sheet)
    grouped: OrderedDict[tuple, dict] = OrderedDict()
    errors: list[str] = []
    piezas = 0
    for row_idx, row in enumerate(sheet.iter_rows(min_row=start, values_only=True), start=start):
        parsed = _parse_row(headers, row)
        if parsed is None:
            continue
        if not parsed["modelo"] and not parsed["conjunto"]:
            errors.append(f"Fila {row_idx}: falta modelo o conjunto.")
            continue
        if not parsed["pieza"]:
            errors.append(f"Fila {row_idx}: falta el nombre de la pieza.")
            continue
        key = (
            parsed.get("id") or "",
            fold(parsed["marca"]),
            fold(parsed["modelo"]),
            fold(parsed["motor"]),
            fold(parsed["anio"]),
            fold(parsed["conjunto"]),
        )
        bucket = grouped.get(key)
        if not bucket:
            bucket = {
                "id": parsed.get("id") or "",
                "marca": parsed["marca"],
                "modelo": parsed["modelo"],
                "motor": parsed["motor"],
                "anio": parsed["anio"],
                "conjunto": parsed["conjunto"],
                "notas": parsed["notas"],
                "venta": parsed["venta"],
                "piezas": [],
            }
            grouped[key] = bucket
        else:
            if parsed["notas"]:
                bucket["notas"] = parsed["notas"]
            if parsed["venta"]:
                bucket["venta"] = parsed["venta"]
        bucket["piezas"].append(
            {
                "nombre": parsed["pieza"],
                "marca": parsed["marca_pieza"],
                "lado": parsed["lado"],
                "codigo": parsed["codigo"],
                "nota": parsed["nota"],
            }
        )
        piezas += 1
    book.close()
    if not grouped:
        raise ExcelError("No encontré filas con auto/conjunto y pieza. Bajá la plantilla y copiá el formato.")
    before_ids = {item["id"] for item in load_all()}
    before_keys = {identity_key(item) for item in load_all()}
    created = 0
    updated = 0
    saved = []
    for ficha in grouped.values():
        key = identity_key(ficha)
        existed = (ficha.get("id") in before_ids and ficha.get("id")) or (
            any(key) and key in before_keys
        )
        item = upsert(ficha)
        saved.append(item)
        if existed:
            updated += 1
        else:
            created += 1
            before_ids.add(item["id"])
            before_keys.add(identity_key(item))
    return {
        "ok": True,
        "fichas": saved,
        "creadas": created,
        "actualizadas": updated,
        "piezas": piezas,
        "errores": errors,
    }


_LEEME = (
    "Una fila = una pieza. Repetí marca/modelo/motor/año/conjunto en cada fila del mismo auto.",
    "Esas filas se agrupan en UNA ficha. Ahí la IA aprende qué llevar y qué ofrecer de más.",
    "Si el cliente pide una sola pieza, el chat usa 'Cómo vender / cruzar' y las otras filas del conjunto.",
    "Marca de la pieza (Sachs, etc.): solo escribí la que venden. La IA no inventa marcas.",
    "Volvé a subir el mismo auto+conjunto para reemplazar esa ficha (no duplica).",
    "Después de cargar, preguntá en el chat como un cliente para verificar.",
)


def _rows_from_fichas(fichas: list[dict]) -> list[dict]:
    rows = []
    for item in fichas:
        piezas = item.get("piezas") or [{}]
        for piece in piezas:
            rows.append(
                {
                    "marca": item.get("marca") or "",
                    "modelo": item.get("modelo") or "",
                    "motor": item.get("motor") or "",
                    "anio": item.get("anio") or "",
                    "conjunto": item.get("conjunto") or "",
                    "notas": item.get("notas") or "",
                    "venta": item.get("venta") or "",
                    "pieza": piece.get("nombre") or "",
                    "marca_pieza": piece.get("marca") or "",
                    "lado": piece.get("lado") or "",
                    "codigo": piece.get("codigo") or "",
                    "nota": piece.get("nota") or "",
                }
            )
    return [row for row in rows if any(str(v).strip() for v in row.values())]


def _write_sheet(sheet: Worksheet, rows: list[dict]) -> None:
    head_font = Font(bold=True, color="F8EFE4")
    head_fill = PatternFill("solid", fgColor="9A3412")
    for col, (key, title) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(1, col, title)
        cell.font = head_font
        cell.fill = head_fill
        sheet.column_dimensions[get_column_letter(col)].width = 22 if col < 8 else 18
    for ridx, row in enumerate(rows, start=2):
        for col, (key, _title) in enumerate(COLUMNS, start=1):
            sheet.cell(ridx, col, row.get(key) or "")
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(len(rows) + 1, 2)}"
    sheet.freeze_panes = "A2"


def _data_sheet(book) -> Worksheet:
    for name in book.sheetnames:
        if fold(name) != "leeme":
            return book[name]
    raise ExcelError("El Excel no tiene una hoja de datos.")


def _find_headers(sheet: Worksheet) -> tuple[dict[int, str], int]:
    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
        mapping = {}
        for col, value in enumerate(row):
            field = _field_from_header(value)
            if field:
                mapping[col] = field
        names = set(mapping.values())
        if "pieza" in names and ("modelo" in names or "conjunto" in names):
            return mapping, idx + 1
    raise ExcelError(
        "No encontré encabezados. Tiene que haber columnas de modelo o conjunto, y pieza."
    )


def _field_from_header(value) -> str:
    token = fold(_cell(value))
    for ch in "()/¿?…":
        token = token.replace(ch, " ")
    token = "_".join(token.replace("/", " ").split())
    if not token:
        return ""
    if "marca" in token and "pieza" in token:
        return "marca_pieza"
    if token.startswith("marca"):
        return "marca"
    if "vender" in token or "cruzar" in token:
        return "venta"
    if token.startswith("pieza") or ("nombre" in token and "pieza" in token):
        return "pieza"
    if token.startswith("notas"):
        return "notas"
    if "nota" in token and "pieza" in token:
        return "nota"
    if token.startswith("nota"):
        return "nota"
    if token.startswith("conjunto") or token.startswith("sistema"):
        return "conjunto"
    if "lado" in token:
        return "lado"
    if "codigo" in token or token == "oem":
        return "codigo"
    if token.startswith("modelo"):
        return "modelo"
    if token.startswith("motor") or token == "cilindrada":
        return "motor"
    if token.startswith("anio") or token.startswith("ano") or token == "year":
        return "anio"
    for field, aliases in _ALIASES.items():
        if token in aliases or token == field:
            return field
    return ""


def _parse_row(headers: dict[int, str], row: tuple) -> dict | None:
    parsed = {key: "" for key, _title in COLUMNS}
    parsed["id"] = ""
    empty = True
    for col, field in headers.items():
        if col >= len(row):
            continue
        text = _cell(row[col])
        if text:
            empty = False
        parsed[field] = text
    if empty:
        return None
    return parsed


def _cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
